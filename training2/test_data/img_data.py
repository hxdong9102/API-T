"""
图片表情包的信息  openpyxl -- 处理速度会相对慢
第三方库 -- 读写  xlrd  -- 只读  xlwt --写入  6万

"""
import openpyxl


excel_path = r"C:\Users\dhx\repository\python\training2\test_data\img_data.xlsx"
wb = openpyxl.load_workbook(excel_path)
# 操作表格 -- 获取打开的表格 获取默认打开
sheet = wb.active
# 获取指定表格
# sheet = wb["Sheet"]
# 锁定某一个单元格
# print(sheet.cell(17,2).value)
# 获取最大行
# 图片路径 文本信息  -- 一组数据
# print(sheet.max_row)  开闭区间 --
# []   -- [[],[],[],[],[],[]]
# 一组组数据 list
img_list = []
# 方法 -- 方便调用
# sheet.cell(1,1).value = "你好"
def get_img_imfo():
    for i in range(1, sheet.max_row + 1):
        # 获取所有的 图片连接
        img_list_info = []
        img_info = sheet.cell(i, 1).value
        img_text = sheet.cell(i, 2).value
        img_list_info.append(img_info)
        img_list_info.append(img_text)
        img_list.append(img_list_info)
    return img_list
#print(get_img_imfo())
