"""
处理excel表格中的信息处理, 得到需要的数据集。 

"""

import openpyxl


# 获取excel表中的信息。
wb = openpyxl.load_workbook(r'D:\Learning-File\API_TEST\course\training3\test_data\dataset.xlsx')
# 操作表格 -- 获取打开的表格 获取默认打开
sheet = wb.active
#print(sheet.max_row)

# 把表格中的数据规范到data_list
data_list = []
def get_data_info():
    for i in range(1, sheet.max_row + 1):
        # 获取所有的 图片连接
        data_list_info = []
        data_info = sheet.cell(i, 1).value.split('+')
        data_info0 = int(data_info[0])
        data_info1 = int(data_info[1])
        data_expre = int(sheet.cell(i, 2).value)
        data_list_info.append(data_info0)
        data_list_info.append(data_info1)
        data_list_info.append(data_expre)
        data_list.append(data_list_info)
    return data_list
    